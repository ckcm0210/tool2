# -*- coding: utf-8 -*-
"""
Created on Wed Jun 25 09:10:23 2025

@author: kccheng
"""

import openpyxl
import os

def _find_matching_sheet(workbook, potential_sheet_name, current_sheet_com_obj=None):
    for ws in workbook.Worksheets:
        if ws.Name == potential_sheet_name:
            return ws
    return None

def _read_external_cell_value(use_openpyxl_var, current_workbook_path, external_file_path, sheet_name, cell_ref):
    if not use_openpyxl_var.get():
        return ""
    if not os.path.isabs(external_file_path):
        base_dir = os.path.dirname(current_workbook_path)
        external_file_path = os.path.join(base_dir, external_file_path)
    try:
        wb = openpyxl.load_workbook(external_file_path, data_only=True)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        val = ws[cell_ref].value
        wb.close()
        return val
    except Exception as e:
        return f"External (openpyxl error: {str(e)})"

def open_external_file(file_path):
    try:
        if os.path.exists(file_path):
            os.startfile(file_path)
    except Exception:
        pass