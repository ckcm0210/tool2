# -*- coding: utf-8 -*-
"""
Created on Wed Jun 25 09:10:53 2025

@author: kccheng
"""

import tkinter as tk
from tkinter import font
from tkinter import filedialog, messagebox, simpledialog, ttk
import win32com.client
import os
import openpyxl
import time
import re
from openpyxl.utils import get_column_letter
from excel_utils import get_referenced_cell_values, parse_excel_address

from worksheet_ui import setup_ui, _set_placeholder, _on_focus_in, _on_mouse_click, _on_focus_out
from worksheet_refresh import refresh_data, reconnect_to_excel, activate_excel_window
from worksheet_tree import apply_filter, sort_column, on_select, on_double_click
from worksheet_export import export_formulas_to_excel, import_and_update_formulas
from worksheet_excel_util import _find_matching_sheet, open_external_file

class WorksheetPane:
    def __init__(self, parent_frame, root_app, pane_name):
        self.parent = parent_frame
        self.root = root_app
        self.pane_name = pane_name
        self.xl = None
        self.workbook = None
        self.worksheet = None
        self.all_formulas = []
        self.cell_addresses = {}
        self.use_openpyxl = tk.BooleanVar(value=True)
        self.show_formula = tk.BooleanVar(value=True)
        self.show_local_link = tk.BooleanVar(value=True)
        self.show_external_link = tk.BooleanVar(value=True)
        self.sort_directions = {col: 1 for col in ("type", "address", "formula", "result", "display_value")}
        self.current_sort_column = None
        self.ui_initialized = False
        self.progress_bar = None
        self.progress_label = None
        self.last_workbook_path = None
        self.last_worksheet_name = None
        self.placeholder_text = "e.g. A, A:A, A:C, Z:A, 10, 10:10, 10:20, 88:17, A1:C3, D40:B5"
        self.placeholder_color = 'grey'
        self.default_fg_color = 'black'
        self.default_font = None
        self.placeholder_font = None
        self.progress_container_frame = tk.Frame(self.parent)
        self.progress_frame = tk.Frame(self.progress_container_frame, bd=2, relief=tk.GROOVE)
        self.progress_bar = ttk.Progressbar(self.progress_frame, orient="horizontal", length=200, mode="determinate")
        self.progress_label = tk.Label(self.progress_frame, text="")
        self.progress_bar.pack(fill=tk.X, expand=True, pady=(5, 2), padx=5)
        self.progress_label.pack(fill=tk.X, pady=(2, 5), padx=5)
        self.progress_frame.pack(fill=tk.X, expand=True, padx=5, pady=5)
        self.setup_ui()

    def setup_ui(self):
        return setup_ui(self)

    def _set_placeholder(self):
        return _set_placeholder(self)

    def _on_focus_in(self, event):
        return _on_focus_in(self, event)

    def _on_mouse_click(self, event):
        return _on_mouse_click(self, event)

    def _on_focus_out(self, event):
        return _on_focus_out(self, event)

    def refresh_data(self, btn, scan_mode='full'):
        return refresh_data(self, btn, scan_mode)

    def reconnect_to_excel(self):
        return reconnect_to_excel(self)

    def activate_excel_window(self):
        return activate_excel_window(self)

    def apply_filter(self, event=None):
        return apply_filter(self, event)

    def sort_column(self, col_id):
        return sort_column(self, col_id)

    def on_select(self, event):
        return on_select(self, event)

    def on_double_click(self, event):
        return on_double_click(self, event)

    def export_formulas_to_excel(self):
        return export_formulas_to_excel(self)

    def import_and_update_formulas(self):
        return import_and_update_formulas(self)

    def summarize_external_links(self):
        return summarize_external_links(self)

    def _find_matching_sheet(self, potential_sheet_name, current_sheet_com_obj=None):
        return _find_matching_sheet(self.workbook, potential_sheet_name, current_sheet_com_obj)

    def _read_external_cell_value(self, current_workbook_path, external_file_path, sheet_name, cell_ref):
        if not self.use_openpyxl.get():
            return "(Reading disabled by user)"
    
        full_external_path_normalized = os.path.normpath(external_file_path)
        if not os.path.exists(full_external_path_normalized):
            return f"External (File Not Found on Disk: {full_external_path_normalized})"
    
        file_extension = os.path.splitext(full_external_path_normalized)[1].lower()
    
        if file_extension in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(full_external_path_normalized, data_only=True, read_only=True)
                found_sheet = None
                for sname in workbook.sheetnames:
                    if sname.lower() == sheet_name.lower():
                        found_sheet = sname
                        break
                if found_sheet:
                    worksheet = workbook[found_sheet]
                    cell_value = worksheet[cell_ref].value
                    workbook.close()
                    return f"External (OpenPyxl): {cell_value if cell_value is not None else 'Empty'}"
                else:
                    workbook.close()
                    return "External (Sheet Not Found in file)"
            except Exception as e:
                return f"External (OpenPyxl Error: {str(e)[:100]})"
    
        elif file_extension == '.xls':
            try:
                import xlrd
                from openpyxl.utils import column_index_from_string
                
                workbook = xlrd.open_workbook(full_external_path_normalized, on_demand=True)
                found_sheet = None
                for sname in workbook.sheet_names():
                    if sname.lower() == sheet_name.lower():
                        found_sheet = sname
                        break
                if found_sheet:
                    worksheet = workbook.sheet_by_name(found_sheet)
                    
                    clean_address = cell_ref.replace('$', '')
                    match = re.match(r"([A-Z]+)([0-9]+)", clean_address)
                    if match:
                        col_str = match.group(1)
                        row_str = match.group(2)
                        col_idx = column_index_from_string(col_str) - 1
                        row_idx = int(row_str) - 1
                        
                        if 0 <= row_idx < worksheet.nrows and 0 <= col_idx < worksheet.ncols:
                            cell_value = worksheet.cell_value(row_idx, col_idx)
                            return f"External (xlrd): {cell_value if cell_value != '' else 'Empty'}"
                        else:
                            return "External (Cell Address Out of Range)"
                    else:
                        return "External (Invalid Cell Address Format)"
                else:
                    return "External (Sheet Not Found in file)"
            except Exception as e:
                return f"External (xlrd Error: {str(e)[:100]})"
    
        return "External (File type not supported for reading)"
    
    def get_referenced_cell_values(self, formula_str, current_sheet_com_obj, current_workbook_path, read_external_cell_value_func, find_matching_sheet_func):
        return get_referenced_cell_values(formula_str, current_sheet_com_obj, current_workbook_path, read_external_cell_value_func, find_matching_sheet_func)

    def parse_excel_address(self, addr):
        return parse_excel_address(addr)
        
    def open_external_file(self, file_path):
        return open_external_file(file_path)
    
    def summarize_external_links(self):
        if not self.result_tree.get_children():
            messagebox.showinfo("No Data", "There are no formulas in the list to summarize.\nPlease scan a worksheet first, or adjust filters.")
            return
    
        summary_window = tk.Toplevel(self.root)
        summary_window.did_replace = False
        summary_window.transient(self.root)
        summary_window.grab_set()
        full_workbook_path = self.workbook.FullName if self.workbook and hasattr(self.workbook, 'FullName') else 'N/A'
        summary_window.title(f"External Link Summary for {full_workbook_path}!{self.worksheet.Name} ({self.pane_name})")
        summary_window.geometry("900x700")
        summary_window.resizable(True, True)
    
        main_frame = ttk.Frame(summary_window, padding=10)
        main_frame.pack(fill='both', expand=True)
        main_frame.rowconfigure(1, weight=1)
        main_frame.columnconfigure(0, weight=1)
    
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=0, column=0, sticky="ew", pady=(0, 5))
    
        formulas_to_summarize = [self.result_tree.item(item, "values") for item in self.result_tree.get_children()]
        
        is_filtered = len(formulas_to_summarize) != len(self.all_formulas) if self.all_formulas else True
    
        heading_text = "External Link Path"
        if is_filtered:
            heading_text += " (Based on user filtered result)"
    
        tree_frame = ttk.LabelFrame(main_frame, text="Found External Links")
        tree_frame.grid(row=1, column=0, sticky="nsew", pady=(0, 10))
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)
    
        summary_tree = ttk.Treeview(tree_frame, columns=("link",), show="headings")
        summary_tree.heading("link", text=heading_text)
        summary_tree.column("link", width=800)
        summary_tree.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=summary_tree.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        summary_tree.configure(yscrollcommand=scrollbar.set)
        
        external_path_pattern = re.compile(
            r"'([^']+\\[^\]]+\.(?:xlsx|xls|xlsm|xlsb)\][^']*?)'", re.IGNORECASE
        )
        unique_full_paths = set()
        try:
            formula_idx = self.tree_columns.index("formula")
        except ValueError:
            messagebox.showerror("Error", "Could not find 'formula' column index.", parent=summary_window)
            summary_window.destroy()
            return
    
        for formula_data in formulas_to_summarize:
            if len(formula_data) > formula_idx:
                formula_content = formula_data[formula_idx]
                matches = external_path_pattern.findall(str(formula_content))
                if matches:
                    unique_full_paths.update(matches)
    
        sorted_full_paths = sorted(list(unique_full_paths))
    
        def show_summary_by_worksheet():
            summary_tree.delete(*summary_tree.get_children())
            for path in sorted_full_paths:
                summary_tree.insert("", "end", values=(path,))
            
            lf_text = "Found External Links (by Worksheet)"
            if is_filtered:
                lf_text += " - Filtered View"
            tree_frame.config(text=lf_text)
    
        def show_summary_by_workbook():
            summary_tree.delete(*summary_tree.get_children())
            unique_workbook_paths = set()
            workbook_only_pattern = re.compile(r"^(.*\\\[[^\]]+\.(?:xlsx|xls|xlsm|xlsb)\])")
            for full_path in sorted_full_paths:
                match = workbook_only_pattern.match(full_path)
                if match:
                    unique_workbook_paths.add(match.group(1))
            
            sorted_workbook_paths = sorted(list(unique_workbook_paths))
            for path in sorted_workbook_paths:
                summary_tree.insert("", "end", values=(path,))
                
            lf_text = "Found External Links (by Workbook)"
            if is_filtered:
                lf_text += " - Filtered View"
            tree_frame.config(text=lf_text)
    
        btn_by_sheet = ttk.Button(button_frame, text="Summarize by Path\\[File]Worksheet", command=show_summary_by_worksheet)
        btn_by_sheet.pack(side='left', padx=5)
        btn_by_workbook = ttk.Button(button_frame, text="Summarize by Path\\[File] only", command=show_summary_by_workbook)
        btn_by_workbook.pack(side='left', padx=5)
    
        replace_frame = ttk.LabelFrame(main_frame, text="Replace Tool", padding=10)
        replace_frame.grid(row=2, column=0, sticky="ew")
        replace_frame.columnconfigure(1, weight=1)
    
        ttk.Label(replace_frame, text="Old Link (Selected):").grid(row=0, column=0, sticky="nw", padx=5, pady=2)
        old_link_var = tk.StringVar(value="<No selection>")
        ttk.Label(replace_frame, textvariable=old_link_var, foreground="maroon", wraplength=700, anchor="w", justify='left').grid(row=0, column=1, columnspan=2, sticky="ew", padx=5, pady=2)
    
        ttk.Label(replace_frame, text="New Link:").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        new_link_entry = ttk.Entry(replace_frame)
        new_link_entry.grid(row=1, column=1, sticky="ew", padx=5, pady=2)
    
        def browse_for_new_link():
            file_path = filedialog.askopenfilename(
                title="Select the new Excel file",
                filetypes=[("Excel Workbooks", "*.xlsx *.xls *.xlsm *.xlsb"), ("All Files", "*.*")],
                parent=summary_window
            )
            if not file_path:
                return
    
            dir_name = os.path.dirname(file_path).replace('/', '\\')
            file_name = os.path.basename(file_path)
            
            new_base_path = f"{dir_name}\\[{file_name}]"
    
            old_link = old_link_var.get()
            worksheet_part = ""
            if old_link != "<No selection>" and ']' in old_link:
                try:
                    worksheet_part = old_link.split(']', 1)[1]
                except IndexError:
                    worksheet_part = ""
            
            final_path = new_base_path + worksheet_part
    
            new_link_entry.delete(0, 'end')
            new_link_entry.insert(0, final_path)
    
        browse_button = ttk.Button(replace_frame, text="...", command=browse_for_new_link, width=4)
        browse_button.grid(row=1, column=2, sticky="w", padx=(2,5), pady=2)           
            
        def on_link_select(event):
            selected_items = summary_tree.selection()
            if selected_items:
                selected_link = summary_tree.item(selected_items[0], "values")[0]
                old_link_var.set(selected_link)
            else:
                old_link_var.set("<No selection>")
        summary_tree.bind("<<TreeviewSelect>>", on_link_select)
            
        def perform_replacement():
            calc_mode_prev = None
            old_link = old_link_var.get()
            newline = "\n"
    
            if old_link == "<No selection>":
                messagebox.showerror(
                    "Replacement Failed - Old Link Not Selected",
                    f"Reason: You have not selected an old link from the list.{newline}{newline}Please select an old link from the 'External Links' list and try again.",
                    parent=summary_window
                )
                return
        
            new_link = new_link_entry.get().strip()
            if not new_link:
                messagebox.showerror(
                    "Replacement Failed - New Link Is Empty",
                    f"Reason: The 'New Link' input field cannot be empty.{newline}{newline}Please enter valid link information in the 'New Link' field.",
                    parent=summary_window
                )
                return
        
            path_match = re.search(r"^(.*\\)?\[([^\]]+)\](.*)$", new_link)
            if not path_match:
                path_example1 = "C:\\path\\[filename.xlsx]Sheetname (with worksheet)"
                path_example2 = "C:\\path\\[filename.xlsx] (f[...]"
                messagebox.showerror(
                    "Replacement Failed - New Link Format Error",
                    f"The format of the 'New Link' is invalid and cannot be recognized.{newline}{newline}"
                    f"Expected format examples:{newline} - {path_example1}{newline} - {path_example2}",
                    parent=summary_window
                )
                return
        
            new_sheet_name_raw = path_match.groups()[2]
            new_sheet_name_cleaned = new_sheet_name_raw.strip("'")
        
            old_link_match = re.search(r"^(.*\\)?\[([^\]]+)\](.*)$", old_link)
        
            if not old_link_match:
                old_sheet_name_raw = ""
            else:
                old_sheet_name_raw = old_link_match.groups()[2]
        
            old_sheet_name_cleaned = old_sheet_name_raw.strip("'")
    
            if old_sheet_name_cleaned != new_sheet_name_cleaned:
                proceed = messagebox.askyesno(
                    "Worksheet Name Mismatch",
                    f"Warning: The worksheet name specified in your old link does not match that in the new link.{newline}{newline}"
                    f"Old Link Worksheet: '{old_sheet_name_cleaned}'{newline}"
                    f"New Link Worksheet: '{new_sheet_name_cleaned}'[...]",
                    parent=summary_window
                )
                if not proceed:
                    messagebox.showinfo("Operation Cancelled", "The replacement operation has been cancelled by the user.", parent=summary_window)
                    return
            
            dir_path, file_name, sheet_name = path_match.groups()
            dir_path = dir_path if dir_path else os.path.dirname(self.workbook.FullName)
            full_file_path = os.path.join(dir_path, file_name)
        
            if not old_sheet_name_cleaned and not new_sheet_name_cleaned:
                old_link_dir_path = ""
                old_link_file_name = ""
        
                if old_link_match:
                    old_link_dir_path, old_link_file_name, _ = old_link_match.groups()
                
                old_full_file_path = os.path.join(old_link_dir_path if old_link_dir_path else os.path.dirname(self.workbook.FullName), old_link_file_name)
        
                if not os.path.exists(old_full_file_path):
                    messagebox.showerror(
                        "Replacement Failed - Old Link File Not Found!",
                        f"Reason: The file pointed to by the old link cannot be found.{newline}{newline}"
                        f"Please check if the path is correct:{newline}'{old_full_file_path}'{newline}{newline}"
                        f"Please ensure the old Excel file exists before per[...]",
                        parent=summary_window
                    )
                    return
                if not os.path.exists(full_file_path):
                    messagebox.showerror(
                        "Replacement Failed - New Link File Not Found!",
                        f"Reason: The file pointed to by the new link cannot be found.{newline}{newline}"
                        f"Please check if the path is correct:{newline}'{full_file_path}'{newline}{newline}"
                        f"Please ensure the new Excel file exists before perform[...]",
                        parent=summary_window
                    )
                    return
        
                old_wb_sheetnames = set()
                new_wb_sheetnames = set()
        
                old_wb = None
                try:
                    old_wb = openpyxl.load_workbook(old_full_file_path, read_only=True)
                    old_wb_sheetnames = set(old_wb.sheetnames)
                except Exception as e:
                    messagebox.showerror(
                        "Replacement Failed - Unable to Read Old File!",
                        f"Reason: An error occurred while trying to read the old link file. This may prevent correct validation of its worksheet names.{newline}{newline}"
                        f"File Path: '{old_full_file_path}'{newline}Error Details:[...]",
                        parent=summary_window
                    )
                    return
                finally:
                    if old_wb:
                        old_wb.close()
                
                new_wb = None
                try:
                    new_wb = openpyxl.load_workbook(full_file_path, read_only=True)
                    new_wb_sheetnames = set(new_wb.sheetnames)
                except Exception as e:
                    messagebox.showerror(
                        "Replacement Failed - Unable to Read New File!",
                        f"Reason: An error occurred while trying to read the new link file. This may prevent correct validation of its worksheet names.{newline}{newline}"
                        f"File Path: '{full_file_path}'{newline}Error Details: {e}[...]",
                        parent=summary_window
                    )
                    return
                finally:
                    if new_wb:
                        new_wb.close()
        
                if old_wb_sheetnames != new_wb_sheetnames:
                    ws_list_old = ", ".join(sorted(list(old_wb_sheetnames))) if old_wb_sheetnames else "None"
                    ws_list_new = ", ".join(sorted(list(new_wb_sheetnames))) if new_wb_sheetnames else "None"
                    messagebox.showerror(
                        "Replacement Failed - Internal Workbook Sheet Names Mismatch!",
                        f"Detected that the worksheet names within the Excel files pointed to by the old and new links do not match.{newline}{newline}"
                        f"Old File Worksheets: {ws_list_old}[...]",
                        parent=summary_window
                    )
                    return
            
            if not os.path.exists(full_file_path):
                messagebox.showerror(
                    "Replacement Failed - New Link File Not Found!",
                    f"Reason: The file pointed to by the new link cannot be found.{newline}{newline}"
                    f"Please check if the path is correct:{newline}'{full_file_path}'{newline}{newline}"
                    f"Please ensure the new Excel file exists before performing [...]",
                    parent=summary_window
                )
                return
    
            try:
                wb = openpyxl.load_workbook(full_file_path, read_only=True)
                cleaned_sheet_name = sheet_name.strip("'")
                if sheet_name and cleaned_sheet_name not in wb.sheetnames:
                    messagebox.showerror(
                        "Replacement Failed - New File Worksheet Not Found!",
                        f'Reason: The worksheet "{cleaned_sheet_name}" specified in the new link was not found in the target file "{file_name}".{newline}{newline}Please check if the worksheet name in the new link is correct, or ensure the worksheet exists in the new Excel file.',
                        parent=summary_window
                    )
                    wb.close()
                    return
                wb.close()
            except Exception as e:
                messagebox.showerror(
                    "Replacement Failed - Unable to Read New File!",
                    f"Reason: An error occurred while trying to read the new link file.{newline}{newline}File Path: '{full_file_path}'{newline}Error Details: {e}{newline}{newline}Please ensure the file is not locked and you have sufficient[...]",
                    parent=summary_window
                )
                return
    
            if not self.worksheet:
                messagebox.showerror(
                    "Replacement Failed - Not Connected to Excel!",
                    f"Reason: The tool is not successfully connected to a live Excel worksheet, thus unable to perform update operations.{newline}{newline}Please ensure you have an Excel file open and the tool is proper[...]",
                    parent=summary_window
                )
                return
            
            affected_cells = []
            formula_idx = self.tree_columns.index("formula")
            address_idx = self.tree_columns.index("address")
    
            current_formulas = [self.result_tree.item(item_id, "values") for item_id in self.result_tree.get_children()]
    
            for item_data in current_formulas:
                if len(item_data) > formula_idx and old_link in str(item_data[formula_idx]):
                    address = item_data[address_idx]
                    formula = item_data[formula_idx]
                    affected_cells.append((address, formula))
    
            if not affected_cells:
                messagebox.showinfo("No Link Found", "The selected old link was not found in any formula in the current view.", parent=summary_window)
                return
    
            try:
                worksheet_name_snapshot = self.worksheet.Name
            except Exception:
                worksheet_name_snapshot = "<Unknown Worksheet>"
            
            confirmation = messagebox.askyesno(
                "Confirm Replacement Operation",
                f"You are about to replace the following link:{newline}{newline}Old Link: {old_link}{newline}{newline}New Link: {new_link}{newline}{newline}This will affect {len(affected_cells)} cells in '{self.worksheet.Name}' worksheet. This ac[...]",
                parent=summary_window
            )
    
            if not confirmation:
                messagebox.showinfo("Operation Cancelled", "The replacement operation has been cancelled by the user.", parent=summary_window)
                return
    
            if self.xl:
                try:
                    calc_mode_prev = self.xl.Calculation
                    self.xl.Calculation = -4135
                except Exception:
                    pass
            
            self.activate_excel_window()
            self.worksheet.Activate()
    
            updated_count = 0
            error_count = 0
            for address, old_formula in affected_cells:
                try:
                    new_formula = old_formula.replace(old_link, new_link)
                    self.worksheet.Range(address).Formula = new_formula
                    updated_count += 1
                except Exception:
                    error_count += 1
    
            for item_id in self.result_tree.get_children():
                values = list(self.result_tree.item(item_id, "values"))
                if len(values) > formula_idx and old_link in str(values[formula_idx]):
                    values[formula_idx] = str(values[formula_idx]).replace(old_link, new_link)
                    self.result_tree.item(item_id, values=values)
    
            messagebox.showinfo(
                "Replacement Complete",
                f"Link replacement operation has finished.{newline}{newline}Successfully updated: {updated_count} cells{newline}Failed to update: {error_count} cells{newline}{newline}Please re-scan the target worksheet to view changes in t[...]",
                parent=summary_window
            )
    
            if self.xl and calc_mode_prev is not None:
                try:
                    self.xl.Calculation = calc_mode_prev
                except Exception:
                    pass
    
            unique_full_paths.clear()
            for formula_data in [self.result_tree.item(item_id, "values") for item_id in self.result_tree.get_children()]:
                if len(formula_data) > formula_idx:
                    formula_content = formula_data[formula_idx]
                    matches = external_path_pattern.findall(str(formula_content))
                    if matches:
                        unique_full_paths.update(matches)
            sorted_full_paths[:] = sorted(list(unique_full_paths))
            show_summary_by_worksheet()
    
            old_link_var.set("<No selection>")
            new_link_entry.delete(0, 'end')
            summary_window.did_replace = True
    
        replace_button = ttk.Button(replace_frame, text="Perform Replacement in Excel", command=perform_replacement)
        replace_button.grid(row=2, column=1, columnspan=2, sticky="e", padx=5, pady=10)
    
        show_summary_by_worksheet()
    
        def on_summary_close():
            if hasattr(summary_window, "did_replace") and summary_window.did_replace:
                app = self.parent.winfo_toplevel().app if hasattr(self.parent.winfo_toplevel(), "app") else None
                if app:
                    if self.pane_name == "Worksheet1":
                        app.scan_left_quick()
                    elif self.pane_name == "Worksheet2":
                        app.scan_right_quick()
            summary_window.destroy()
        summary_window.protocol("WM_DELETE_WINDOW", on_summary_close)