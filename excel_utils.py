# -*- coding: utf-8 -*-
"""
Created on Wed Jun 25 09:07:45 2025

@author: kccheng
"""

import re
import os
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

def classify_formula_type(formula):
    formula_str = str(formula)
    external_link_pattern = re.compile(r"\[([^\]]+?\.(?:xlsx|xls|xlsm|xlsb))\]", re.IGNORECASE)
    if external_link_pattern.search(formula_str):
        return 'external link'
    if '!' in formula_str:
        return 'local link'
    if formula_str.startswith('='):
        return 'formula'
    return 'formula'

def is_external_link_regex_match(formula_str):
    external_link_pattern = re.compile(r"\[([^\]]+?\.(?:xlsx|xls|xlsm|xlsb))\]", re.IGNORECASE)
    return bool(external_link_pattern.search(formula_str))

def parse_excel_address(addr):
    addr = addr.replace('$', '').strip().upper()

    if not addr:
        raise ValueError("Address input cannot be empty.")

    if re.fullmatch(r"^[0-9]+(:[0-9]+)?$", addr):
        parts = list(map(int, addr.split(':')))
        start, end = (parts[0], parts[0]) if len(parts) == 1 else (parts[0], parts[1])
        if start > end:
            start, end = end, start
        return ('row_range', f"{start}:{end}")

    if re.fullmatch(r"^[A-Z]+(:[A-Z]+)?$", addr):
        parts = addr.split(':')
        start_col, end_col = (parts[0], parts[0]) if len(parts) == 1 else (parts[0], parts[1])
        start_idx = column_index_from_string(start_col)
        end_idx = column_index_from_string(end_col)
        if start_idx > end_idx:
            start_idx, end_idx = end_idx, start_idx
        start_col_sorted = get_column_letter(start_idx)
        end_col_sorted = get_column_letter(end_idx)
        return ('col_range', f"{start_col_sorted}:{end_col_sorted}")

    if re.fullmatch(r"^[A-Z]+[0-9]+$", addr):
        return ('cell', addr)
        
    m = re.fullmatch(r"^([A-Z]+[0-9]+):([A-Z]+[0-9]+)$", addr)
    if m:
        c1, c2 = m.groups()
        c1_col_str, c1_row_str = re.match(r"([A-Z]+)([0-9]+)", c1).groups()
        c2_col_str, c2_row_str = re.match(r"([A-Z]+)([0-9]+)", c2).groups()
        
        c1_col = column_index_from_string(c1_col_str)
        c2_col = column_index_from_string(c2_col_str)
        c1_row = int(c1_row_str)
        c2_row = int(c2_row_str)

        start_col_idx = min(c1_col, c2_col)
        end_col_idx = max(c1_col, c2_col)
        start_row = min(c1_row, c2_row)
        end_row = max(c1_row, c2_row)
        
        start_cell = f"{get_column_letter(start_col_idx)}{start_row}"
        end_cell = f"{get_column_letter(end_col_idx)}{end_row}"
        
        return ('range', f"{start_cell}:{end_cell}")

    raise ValueError(f"Invalid address format: '{addr}'")


def get_referenced_cell_values(
    formula_str, 
    current_sheet_com_obj, 
    current_workbook_path,
    read_external_cell_value_func,
    find_matching_sheet_func
):
    referenced_data = {}
    processed_spans = []

    def is_span_processed(start, end):
        for p_start, p_end in processed_spans:
            if start < p_end and end > p_start:
                return True
        return False

    def add_processed_span(start, end):
        processed_spans.append((start, end))

    patterns = [
        (
            'external',
            re.compile(
                r"'?((?:[a-zA-Z]:\\)?[^']*)\[([^\]]+\.(?:xlsx|xls|xlsm|xlsb))\]([^']*)'?\s*!\s*(\$?[A-Z]{1,3}\$?\d{1,7}(?::\$?[A-Z]{1,3}\$?\d{1,7})?)",
                re.IGNORECASE
            )
        ),
        (
            'local_quoted',
            re.compile(
                r"'([^']+)'!(\$?[A-Z]{1,3}\$?\d{1,7}(?::\$?[A-Z]{1,3}\$?\d{1,7})?)",
                re.IGNORECASE
            )
        ),
        (
            'local_unquoted',
            re.compile(
                r"([a-zA-Z0-9_\u4e00-\u9fa5][a-zA-Z0-9_\s\.\u4e00-\u9fa5]{0,30})!(\$?[A-Z]{1,3}\$?\d{1,7}(?::\$?[A-Z]{1,3}\$?\d{1,7})?)",
                re.IGNORECASE
            )
        ),
        (
            'current_range',
            re.compile(
                r"(?<![!'\[\]a-zA-Z0-9_\u4e00-\u9fa5])(\$?[A-Z]{1,3}\$?\d{1,7}:\s*\$?[A-Z]{1,3}\$?\d{1,7})(?![a-zA-Z0-9_])",
                re.IGNORECASE
            )
        ),
        (
            'current_single',
            re.compile(
                r"(?<![!'\[\]a-zA-Z0-9_\u4e00-\u9fa5])(\$?[A-Z]{1,3}\$?\d{1,7})(?![a-zA-Z0-9_:\(])",
                re.IGNORECASE
            )
        )
    ]

    all_matches = []
    for p_type, pattern in patterns:
        for match in pattern.finditer(formula_str):
            all_matches.append({'type': p_type, 'match': match, 'span': match.span()})

    all_matches.sort(key=lambda x: (x['span'][0], x['span'][1] - x['span'][0]))

    for item in all_matches:
        match = item['match']
        m_type = item['type']
        start, end = item['span']

        if is_span_processed(start, end):
            continue

        try:
            if m_type == 'external':
                dir_path, file_name, sheet_name, cell_ref = match.groups()
                sheet_name = sheet_name.strip("'")
                full_file_path = os.path.join(dir_path, file_name)
                display_ref = f"[{os.path.basename(full_file_path)}]{sheet_name}!{cell_ref.replace('$', '')}"
                display_ref_with_path = f"{full_file_path}|{display_ref}"

                if ':' in cell_ref:
                    value = "(Range Reference)"
                else:
                    value = read_external_cell_value_func(
                        current_workbook_path, full_file_path, sheet_name, cell_ref.replace('$', '')
                    )
                if display_ref_with_path not in referenced_data:
                    referenced_data[display_ref_with_path] = value

            elif m_type in ('local_quoted', 'local_unquoted'):
                sheet_name, cell_ref = match.groups()
                sheet_name = sheet_name.strip("'")
                
                if sheet_name.lower().endswith(('.xlsx', '.xls', '.xlsm', '.xlsb')):
                    continue

                display_ref = f"{sheet_name}!{cell_ref.replace('$', '')}"
                
                if ':' in cell_ref:
                    value = "(Range Reference)"
                else:
                    target_sheet = find_matching_sheet_func(sheet_name, current_sheet_com_obj)
                    if target_sheet:
                        cell_val = target_sheet.Range(cell_ref).Value
                        value = f"Local: {cell_val if cell_val is not None else 'Empty'}"
                    else:
                        value = f"Local (Sheet '{sheet_name}' Not Found)"
                
                if display_ref not in referenced_data:
                    referenced_data[display_ref] = value

            elif m_type in ('current_range', 'current_single'):
                cell_ref = match.group(1)
                display_ref = f"{current_sheet_com_obj.Name}!{cell_ref.replace('$', '')}"

                if ':' in cell_ref:
                    value = "(Range Reference)"
                else:
                    cell_val = current_sheet_com_obj.Range(cell_ref).Value
                    value = f"Current: {cell_val if cell_val is not None else 'Empty'}"
                
                if display_ref not in referenced_data:
                    referenced_data[display_ref] = value

            add_processed_span(start, end)
        except Exception as e:
            print(f"ERROR: Could not process reference from match '{match.group(0)}': {e}")

    return referenced_data

def calculate_similarity(str1, str2):
    len1, len2 = len(str1), len(str2)
    if len1 == 0 or len2 == 0:
        return 0.0
    dp = [[0] * (len2 + 1) for _ in range(len1 + 1)]
    for i in range(len1 + 1):
        dp[i][0] = i
    for j in range(len2 + 1):
        dp[0][j] = j
    for i in range(1, len1 + 1):
        for j in range(1, len2 + 1):
            if str1[i-1] == str2[j-1]:
                dp[i][j] = dp[i-1][j-1]
            else:
                dp[i][j] = min(dp[i-1][j], dp[i][j-1], dp[i-1][j-1]) + 1
    edit_distance = dp[len1][len2]
    max_len = max(len1, len2)
    similarity = 1.0 - (edit_distance / max_len)
    return similarity

def parse_external_path_and_sheet(path_and_sheet):
    if '[' in path_and_sheet and ']' in path_and_sheet:
        file_and_sheet = path_and_sheet.split('[')[1]
        if ']' in file_and_sheet:
            file_name = file_and_sheet.split(']')[0]
            sheet_name_part = file_and_sheet.split(']')[1]
            if sheet_name_part.startswith("'"):
                sheet_name = sheet_name_part.lstrip("'")
            else:
                sheet_name = sheet_name_part.strip('!')
        else:
            file_name = file_and_sheet
            sheet_name = ''
    else:
        file_name = ''
        sheet_name = path_and_sheet.strip('!')
    return file_name, sheet_name

def read_external_cell_value(current_workbook_path, external_file_full_path, external_sheet_name, cell_address):
    import os
    full_external_path_normalized = os.path.normpath(external_file_full_path)
    if not os.path.exists(full_external_path_normalized):
        return f"External (File Not Found on Disk: {full_external_path_normalized})"
    file_extension = os.path.splitext(full_external_path_normalized)[1].lower()
    if file_extension in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(full_external_path_normalized, data_only=True, read_only=True)
            found_sheet = None
            for sname in workbook.sheetnames:
                if sname.lower() == external_sheet_name.lower():
                    found_sheet = sname
                    break
            if found_sheet:
                worksheet = workbook[found_sheet]
                cell_value = worksheet[cell_address].value
                workbook.close()
                return f"External (OpenPyxl): {cell_value if cell_value is not None else 'Empty'}"
            else:
                workbook.close()
                return "External (Sheet Not Found in file)"
        except Exception as e:
            return f"External (OpenPyxl Error: {str(e)[:100]})"
    if file_extension == '.xls':
        try:
            import xlrd
            workbook = xlrd.open_workbook(full_external_path_normalized, on_demand=True)
            found_sheet = None
            for sname in workbook.sheet_names():
                if sname.lower() == external_sheet_name.lower():
                    found_sheet = sname
                    break
            if found_sheet:
                worksheet = workbook.sheet_by_name(found_sheet)
                import re
                m = re.match(r'^([A-Z]+)([0-9]+)$', cell_address.replace('$', ''))
                if m:
                    col_letters, row_str = m.groups()
                    col_idx = 0
                    for i, c in enumerate(reversed(col_letters)):
                        col_idx += (ord(c.upper()) - ord('A') + 1) * (26 ** i)
                    col_idx -= 1
                    row_idx = int(row_str) - 1
                    if 0 <= row_idx < worksheet.nrows and 0 <= col_idx < worksheet.ncols:
                        cell_value = worksheet.cell_value(row_idx, col_idx)
                        return f"External (xlrd): {cell_value if cell_value != '' else 'Empty'}"
                    else:
                        return "External (Cell Address Out of Range)"
                else:
                    return "External (Invalid Cell Address Format)"
            else:
                return "External (Sheet Not Found in file)"
        except Exception as e:
            return f"External (xlrd Error: {str(e)[:100]})"
    return "External (Live reading for this file type is disabled)"

def find_matching_sheet(workbook, sheet_name):
    try:
        for ws in workbook.Worksheets:
            if ws.Name == sheet_name:
                return ws
    except Exception as e:
        print(f"ERROR: Failed to get worksheet names: {e}")
    return None

def get_sheet_by_name(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Worksheet '{sheet_name}' not found in this workbook!")
    return wb[sheet_name]